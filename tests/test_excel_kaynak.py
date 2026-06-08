"""Excel 'Kaynak' (Dijital/OCR/XML) sütunu."""

from openpyxl import load_workbook

from excel_utils import excel_olustur, mevcut_verileri_oku


def _satir(**ek):
    s = {
        "fatura_no":            "GIB2024123456789",
        "vergiler_dahil_tutar": 100.0,
        "dosya_yolu":           "C:\\faturalar\\a.pdf",
    }
    s.update(ek)
    return s


def test_kaynak_basligi_eklenir(tmp_path):
    cikti = str(tmp_path / "f.xlsx")
    excel_olustur([_satir(_teknik_bilgi="OCR")], cikti)
    ws = load_workbook(cikti).active
    basliklar = [c.value for c in ws[1]]
    assert "Kaynak" in basliklar


def test_kaynak_roundtrip_korunur(tmp_path):
    cikti = str(tmp_path / "f.xlsx")
    excel_olustur([_satir(_teknik_bilgi="OCR")], cikti)
    okunan, islenenmis = mevcut_verileri_oku(cikti)
    assert okunan[0]["_teknik_bilgi"] == "OCR"
    assert "a.pdf" in islenenmis


def test_dijital_kaynak_yazilir(tmp_path):
    cikti = str(tmp_path / "f.xlsx")
    excel_olustur([_satir(_teknik_bilgi="Dijital")], cikti)
    okunan, _ = mevcut_verileri_oku(cikti)
    assert okunan[0]["_teknik_bilgi"] == "Dijital"


def test_xml_dosyasi_xml_etiketi_alir(tmp_path):
    # _teknik_bilgi verilmeyen XML dosyası "XML" olarak etiketlenir
    cikti = str(tmp_path / "f.xlsx")
    excel_olustur([_satir(dosya_yolu="C:\\faturalar\\b.xml", _teknik_bilgi=None)], cikti)
    okunan, _ = mevcut_verileri_oku(cikti)
    assert okunan[0]["_teknik_bilgi"] == "XML"

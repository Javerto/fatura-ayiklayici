"""Excel yazma/okuma güvenliği — veri kaybına yol açan yolların testleri."""
import datetime
import os

import pytest
from openpyxl import load_workbook

from excel_utils import ExcelHatasi, excel_olustur, mevcut_verileri_oku


def _satir(no="GIB2024000000101"):
    return {"fatura_no": no, "sirket_adi": "ACME A.Ş.", "vkn": "1234567890",
            "vergiler_dahil_tutar": 1180.0, "kdv_haric_tutar": 1000.0,
            "para_birimi": "TRY", "fatura_tarihi": datetime.datetime(2024, 1, 15),
            "dosya_yolu": r"C:\x\a.pdf", "_teknik_bilgi": "Dijital"}


def test_ozet_sekmesi_aktifken_faturalar_okunur(tmp_path):
    """Kullanıcı Excel'i Özet sekmesinde kaydederse `wb.active` oraya işaret eder.

    Özet hücreleri fatura sanılırsa `islenenmis` boşalır (her PDF yeniden
    işlenir) ve `mevcut` çöple dolar — bir sonraki onay tüm geçmişi çöple
    değiştirir. Bu testin kırılması veri kaybı demektir.
    """
    yol = str(tmp_path / "faturalar.xlsx")
    excel_olustur([_satir()], yol)

    wb = load_workbook(yol)
    assert "Özet" in wb.sheetnames
    wb.active = wb.sheetnames.index("Özet")
    wb.save(yol)

    satirlar, islenenmis = mevcut_verileri_oku(yol)
    assert [s["fatura_no"] for s in satirlar] == ["GIB2024000000101"]
    assert islenenmis == {"a.pdf"}


def test_okunamayan_dosya_hata_firlatir(tmp_path):
    """Sessizce ([], set()) dönerse bir sonraki onay dosyayı sıfırdan yazar."""
    yol = tmp_path / "bozuk.xlsx"
    yol.write_bytes(b"bu bir xlsx degil")
    with pytest.raises(ExcelHatasi):
        mevcut_verileri_oku(str(yol))


def test_kaydetme_hatasinda_mevcut_dosya_bozulmaz(tmp_path, monkeypatch):
    """Yazma yarıda kesilirse eski dosya olduğu gibi kalmalı, .tmp temizlenmeli."""
    yol = str(tmp_path / "faturalar.xlsx")
    excel_olustur([_satir("GIB2024000000101")], yol)
    onceki = open(yol, "rb").read()

    from openpyxl.workbook.workbook import Workbook

    def patlayan_save(self, dosya):
        raise OSError("disk dolu")

    monkeypatch.setattr(Workbook, "save", patlayan_save)
    with pytest.raises(ExcelHatasi):
        excel_olustur([_satir("GIB2024000000102")], yol)

    assert open(yol, "rb").read() == onceki
    assert not os.path.exists(yol + ".tmp")


def test_kilitli_dosya_excel_hatasina_donusur(tmp_path, monkeypatch):
    """PermissionError da, diğer OSError'lar da ExcelHatasi olmalı.

    Aksi hâlde ham hata api.py'nin `except ExcelHatasi`'ından kaçıp JS'e
    sızıyor ve gözden geçirme ekranındaki 'Onayla' butonu ölü kalıyor.
    """
    from openpyxl.workbook.workbook import Workbook
    monkeypatch.setattr(Workbook, "save",
                        lambda self, d: (_ for _ in ()).throw(PermissionError()))
    with pytest.raises(ExcelHatasi, match="Excel'de açık"):
        excel_olustur([_satir()], str(tmp_path / "f.xlsx"))

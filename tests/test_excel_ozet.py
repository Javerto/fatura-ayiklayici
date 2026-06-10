"""Excel 'Özet' sayfası entegrasyon testleri."""
from datetime import datetime

from openpyxl import load_workbook

from excel_utils import excel_olustur, mevcut_verileri_oku


def _satirlar(tmp_path):
    return [
        {"fatura_no": "GIB2024123456789", "sirket_adi": "ACME",
         "fatura_tarihi": datetime(2024, 3, 15), "vergiler_dahil_tutar": 120.0,
         "kdv_haric_tutar": 100.0, "para_birimi": "TL", "vkn": "1234567890",
         "dosya_yolu": str(tmp_path / "a.pdf"), "_teknik_bilgi": "Dijital"},
        {"fatura_no": "ABC2024000000001", "sirket_adi": "Beta",
         "fatura_tarihi": datetime(2024, 4, 1), "vergiler_dahil_tutar": 50.0,
         "kdv_haric_tutar": 40.0, "para_birimi": "EUR", "vkn": "9876543210",
         "dosya_yolu": str(tmp_path / "b.xml"), "_teknik_bilgi": ""},
    ]


def test_ozet_sayfasi_olusur_ve_ana_sayfa_aktif_kalir(tmp_path):
    cikti = str(tmp_path / "c.xlsx")
    excel_olustur(_satirlar(tmp_path), cikti)
    wb = load_workbook(cikti)
    assert wb.sheetnames == ["Faturalar", "Özet"]
    assert wb.active.title == "Faturalar"
    ws = wb["Özet"]
    hucre_metinleri = [str(c.value) for row in ws.iter_rows()
                      for c in row if c.value is not None]
    assert "GENEL" in hucre_metinleri
    assert "AYLIK" in hucre_metinleri
    assert "ŞİRKET" in hucre_metinleri
    assert "Toplam Tutar (TL)" in hucre_metinleri
    assert "Toplam Tutar (EUR)" in hucre_metinleri
    wb.close()


def test_ozet_sayfali_dosyada_roundtrip_bozulmaz(tmp_path):
    cikti = str(tmp_path / "c.xlsx")
    excel_olustur(_satirlar(tmp_path), cikti)
    satirlar, islenmis = mevcut_verileri_oku(cikti)
    assert len(satirlar) == 2
    assert satirlar[0]["fatura_no"] == "GIB2024123456789"
    assert "a.pdf" in islenmis and "b.xml" in islenmis


def test_ikinci_yazimda_ozet_tek_kalir(tmp_path):
    # Dosya seviyesinde davranış güvencesi: excel_olustur her çağrıda taze
    # Workbook kurduğu için üst üste yazımlarda Özet sayfası çoğalamaz/bayatlamaz.
    cikti = str(tmp_path / "c.xlsx")
    excel_olustur(_satirlar(tmp_path), cikti)
    excel_olustur(_satirlar(tmp_path), cikti)
    wb = load_workbook(cikti)
    assert wb.sheetnames.count("Özet") == 1
    wb.close()


def test_tutarsiz_satir_kirilimda_ad_ve_adetle_gorunur(tmp_path):
    # Tutarı olmayan satır (vergiler_dahil_tutar=None) kırılım tablosunda
    # ad + adet ile yer almalı (pb=None dalı), tutar hücresi boş kalmalı.
    satir = {"fatura_no": "GIB2024123456780", "sirket_adi": "Tutarsiz AS",
             "fatura_tarihi": datetime(2024, 5, 1),
             "vergiler_dahil_tutar": None, "kdv_haric_tutar": None,
             "para_birimi": "TL", "vkn": "1234567890",
             "dosya_yolu": str(tmp_path / "t.pdf"), "_teknik_bilgi": "OCR"}
    cikti = str(tmp_path / "c.xlsx")
    excel_olustur([satir], cikti)
    ws = load_workbook(cikti)["Özet"]
    hucreler = [str(c.value) for row in ws.iter_rows()
                for c in row if c.value is not None]
    assert "Tutarsiz AS" in hucreler          # şirket kırılımında görünür
    assert "2024-05" in hucreler              # aylık kırılımda görünür


def test_bos_listede_ozet_sayfasi_yine_olusur(tmp_path):
    cikti = str(tmp_path / "c.xlsx")
    excel_olustur([], cikti)
    wb = load_workbook(cikti)
    assert "Özet" in wb.sheetnames
    wb.close()

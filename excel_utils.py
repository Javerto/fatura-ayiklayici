"""
Excel okuma ve yazma işlemleri.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from hatalar import ExcelHatasi
from fatura import ALANLAR, SUTUN
from fatura import kaynak as fatura_kaynak   # yerel `kaynak` adlarıyla çakışmasın
from ozet import ozet_hesapla

# Başlık ve genişlikler sütun sırasına göre: tablo fatura.ALANLAR'da.
SUTUN_BILGISI = sorted((a for a in ALANLAR if a.baslik), key=lambda a: a.sutun)
# Satırdan doğrudan yazılan/okunan alanlar (formül ve dosya sütunları hariç).
VERI_ALANLARI = [a for a in ALANLAR if a.etiket]


def _bos(alan):
    """Satırda alan yoksa hücreye yazılacak değer: sayıda boş hücre, metinde ''."""
    if alan.anahtar == "para_birimi":
        return "TL"
    return None if alan.tip == "sayi" else ""


def excel_guvenli(deger):
    """Excel'de izin verilmeyen karakterleri temizler."""
    if not isinstance(deger, str):
        return deger
    return ILLEGAL_CHARACTERS_RE.sub("", deger)


def dosya_url(dosya_yolu: str) -> str:
    """Dosya yolunu geçerli file:// URL'e çevirir. UNC path'leri de destekler."""
    p = dosya_yolu.replace("\\", "/")
    if p.startswith("//"):
        # UNC path: \\sunucu\paylasim → file://sunucu/paylasim/...
        return "file:" + p
    else:
        # Yerel path: C:\klasor → file:///C:/klasor/...
        return "file:///" + p


def url_dosya(url: str) -> str:
    """file:// URL'inden dosya yoluna çevirir."""
    if url.startswith("file:////"):
        # Bozuk UNC (eski format): file:////sunucu → \\sunucu
        return url[8:].replace("/", "\\")
    elif url.startswith("file://") and not url.startswith("file:///"):
        # Doğru UNC: file://sunucu → \\sunucu
        return "\\\\" + url[7:].replace("/", "\\")
    else:
        # Yerel: file:///C:/... → C:\...
        return url.replace("file:///", "").replace("/", "\\")


def mevcut_verileri_oku(cikti: str) -> tuple[list[dict], set[str]]:
    """Mevcut Excel'deki faturaları ve işlenmiş dosya adlarını okur."""
    if not os.path.exists(cikti):
        return [], set()
    try:
        wb = load_workbook(cikti, data_only=True)
    except Exception as e:
        # Okunamayan dosyayı "kayıt yok" sayarsak bir sonraki onay onu
        # sıfırdan yazar ve birikmiş tüm geçmiş gider. Okuyamıyorsak durmalıyız.
        raise ExcelHatasi(
            f"'{os.path.basename(cikti)}' okunamadı: {e}\n"
            "Dosya bozuk veya başka bir programda açık olabilir. "
            "Üzerine yazmamak için işlem durduruldu.") from e

    # wb.active, dosyanın en son hangi sekmede kaydedildiğine bakar; kullanıcı
    # Özet sekmesinde bırakıp kaydettiyse oranın hücreleri fatura sanılır.
    ws = wb["Faturalar"] if "Faturalar" in wb.sheetnames else wb.worksheets[0]
    satirlar, islenenmis = [], set()

    for row in ws.iter_rows(min_row=2, max_col=SUTUN["kaynak"]):
        b_val = row[SUTUN["fatura_no"] - 1].value
        if not b_val or str(b_val).strip().upper() == "TOPLAM":
            continue
        kaynak_val = row[SUTUN["kaynak"] - 1].value
        s = {a.anahtar: row[a.sutun - 1].value for a in VERI_ALANLARI}
        s["_teknik_bilgi"] = str(kaynak_val).strip() if kaynak_val else ""
        n_cell     = row[SUTUN["dosya"] - 1]
        gizli_cell = row[SUTUN["dosya_yolu_gizli"] - 1]
        dosya_yolu = ""
        # O sütunu: tam dosya yolu — hyperlink bozulsa bile güvenilir kaynak
        gizli_val = str(gizli_cell.value).strip() if gizli_cell.value else ""
        if gizli_val.lower().endswith((".pdf", ".xml")):
            dosya_yolu = gizli_val
            islenenmis.add(os.path.basename(gizli_val).lower())
        else:
            # Geriye dönük uyumluluk: eski format — hyperlink'ten oku
            if n_cell.hyperlink and n_cell.hyperlink.target:
                dosya_yolu = url_dosya(n_cell.hyperlink.target)
                islenenmis.add(os.path.basename(dosya_yolu).lower())
            else:
                n_val = str(n_cell.value).strip() if n_cell.value else ""
                if n_val.lower().endswith((".pdf", ".xml")):
                    islenenmis.add(n_val.lower())
        s["dosya_yolu"] = dosya_yolu
        satirlar.append(s)

    wb.close()
    return satirlar, islenenmis


def excel_olustur(satirlar: list, cikti: str):
    """Fatura listesinden biçimlendirilmiş Excel dosyası oluşturur."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturalar"

    baslik_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    baslik_fill = PatternFill("solid", start_color="2F5496")
    orta        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ince        = Side(style="thin", color="CCCCCC")
    kenar       = Border(left=ince, right=ince, top=ince, bottom=ince)
    f10         = Font(name="Arial", size=10)
    flink       = Font(name="Arial", size=10, color="0563C1", underline="single")
    fbold       = Font(name="Arial", bold=True, size=10)
    num_fmt     = "#,##0.00"
    zebra_fill  = PatternFill("solid", start_color="EEF2F9")
    toplam_fill = PatternFill("solid", start_color="D9E1F2")

    for a in SUTUN_BILGISI:
        c = ws.cell(row=1, column=a.sutun, value=a.baslik)
        c.font, c.fill, c.alignment, c.border = baslik_font, baslik_fill, orta, kenar
    ws.row_dimensions[1].height = 30

    gizli_harf = get_column_letter(SUTUN["dosya_yolu_gizli"])
    fn_hrf  = get_column_letter(SUTUN["fatura_no"])
    ft_hrf  = get_column_letter(SUTUN["fatura_tarihi"])
    vdt_hrf = get_column_letter(SUTUN["vergiler_dahil_tutar"])
    kdv_hrf = get_column_letter(SUTUN["kdv_haric_tutar"])
    tm_hrf  = get_column_letter(SUTUN["toplam_miktar"])
    vt_hrf  = get_column_letter(SUTUN["vergi_tutari"])

    for ri, s in enumerate(satirlar, 2):
        zebra = zebra_fill if ri % 2 == 0 else None
        ws.cell(row=ri, column=SUTUN["ft_formul"],
                value=f"=+{fn_hrf}{ri}&{ft_hrf}{ri}")

        veri_map = {a.sutun: s.get(a.anahtar, _bos(a)) for a in VERI_ALANLARI}
        for col, val in veri_map.items():
            c = ws.cell(row=ri, column=col, value=excel_guvenli(val))
            c.font, c.border = f10, kenar
            c.alignment = Alignment(vertical="center")
            if col in (SUTUN["vergiler_dahil_tutar"], SUTUN["kdv_haric_tutar"]) and isinstance(val, (int, float)):
                c.number_format = num_fmt
            if col == SUTUN["fatura_tarihi"] and hasattr(val, "strftime"):
                c.number_format = "DD.MM.YYYY"
            if zebra:
                c.fill = zebra

        jc = ws.cell(row=ri, column=SUTUN["vergi_tutari"],
                     value=f"=+{vdt_hrf}{ri}-{kdv_hrf}{ri}")
        jc.font, jc.border, jc.number_format = f10, kenar, num_fmt
        jc.alignment = Alignment(vertical="center")
        if zebra:
            jc.fill = zebra

        ac = ws.cell(row=ri, column=SUTUN["ft_formul"])
        ac.font, ac.border = f10, kenar
        if zebra:
            ac.fill = zebra

        lh = ws.cell(row=ri, column=SUTUN["dosya"])
        dosya_yolu = s.get("dosya_yolu", "")
        if dosya_yolu.lower().endswith(".pdf"):
            lh.value = (f'=HYPERLINK("file:///"&SUBSTITUTE({gizli_harf}{ri},'
                        f'"\\\\","/"),"Faturayı Aç")')
            lh.font = flink
        elif dosya_yolu.lower().endswith(".xml"):
            lh.value = "XML"
            lh.font = f10
        else:
            lh.font = f10
        lh.border = kenar
        lh.alignment = Alignment(horizontal="center", vertical="center")
        if zebra:
            lh.fill = zebra

        # Gizli sütun: tam dosya yolu — hyperlink bozulsa bile kod buradan okur
        gc = ws.cell(row=ri, column=SUTUN["dosya_yolu_gizli"], value=dosya_yolu)
        gc.font = Font(name="Arial", size=1, color="FFFFFF")  # beyaz, 1pt — görünmez
        gc.alignment = Alignment(horizontal="center", vertical="center")

        # Kaynak sütunu: Dijital / OCR (PDF) veya XML
        kc = ws.cell(row=ri, column=SUTUN["kaynak"], value=fatura_kaynak(s))
        kc.font, kc.border = f10, kenar
        kc.alignment = Alignment(horizontal="center", vertical="center")
        if zebra:
            kc.fill = zebra

    for a in SUTUN_BILGISI:
        ws.column_dimensions[get_column_letter(a.sutun)].width = a.genislik
    ws.column_dimensions[gizli_harf].hidden = True

    sr = len(satirlar) + 2
    ws.cell(row=sr, column=SUTUN["fatura_no"], value="TOPLAM").font = fbold
    for col, fml in [
        (SUTUN["toplam_miktar"],        f"=SUM({tm_hrf}2:{tm_hrf}{sr-1})"),
        (SUTUN["vergiler_dahil_tutar"], f"=SUM({vdt_hrf}2:{vdt_hrf}{sr-1})"),
        (SUTUN["kdv_haric_tutar"],      f"=SUM({kdv_hrf}2:{kdv_hrf}{sr-1})"),
        (SUTUN["vergi_tutari"],         f"=SUM({vt_hrf}2:{vt_hrf}{sr-1})"),
    ]:
        c = ws.cell(row=sr, column=col, value=fml)
        c.font = fbold
        if col in (SUTUN["vergiler_dahil_tutar"], SUTUN["kdv_haric_tutar"], SUTUN["vergi_tutari"]):
            c.number_format = num_fmt
    for col in range(1, 15):
        c = ws.cell(row=sr, column=col)
        c.border, c.fill = kenar, toplam_fill
    tc = ws.cell(row=sr, column=SUTUN["kaynak"])
    tc.border, tc.fill = kenar, toplam_fill

    _ozet_sayfasi_yaz(wb, satirlar)

    _guvenli_kaydet(wb, cikti)


def _guvenli_kaydet(wb, cikti: str):
    """Önce geçici dosyaya yazar, sonra yerine taşır.

    Doğrudan `wb.save(cikti)` hedefi anında sıfırlıyor; yazma yarıda kesilirse
    (disk dolu, ağ sürücüsü koptu, süreç öldürüldü) geriye yarım bir dosya
    kalıyor ve birikmiş tüm fatura geçmişi gidiyordu. `os.replace` Windows'ta
    atomiktir: ya eski dosya ya yenisi durur, arası yok.
    """
    gecici = cikti + ".tmp"
    try:
        wb.save(gecici)
        os.replace(gecici, cikti)
    except PermissionError as e:
        _sil(gecici)
        raise ExcelHatasi(
            f"'{os.path.basename(cikti)}' kaydedilemedi.\n"
            "Dosya Excel'de açık olabilir. Kapatıp tekrar deneyin.") from e
    except OSError as e:
        _sil(gecici)
        raise ExcelHatasi(
            f"'{os.path.basename(cikti)}' kaydedilemedi: {e}\n"
            "Mevcut dosyaya dokunulmadı.") from e


def _sil(yol: str):
    try:
        os.remove(yol)
    except OSError:
        pass


def _ozet_sayfasi_yaz(wb, satirlar: list):
    """Çalışma kitabına 'Özet' sayfası ekler (genel/aylık/şirket blokları)."""
    o = ozet_hesapla(satirlar)
    ws = wb.create_sheet("Özet")

    baslik_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    baslik_fill = PatternFill("solid", start_color="2F5496")
    fbold = Font(name="Arial", bold=True, size=10)
    f10 = Font(name="Arial", size=10)
    num_fmt = "#,##0.00"

    r = 1

    def blok_baslik(metin):
        nonlocal r
        for col in range(1, 5):
            c = ws.cell(row=r, column=col)
            c.fill = baslik_fill
        c = ws.cell(row=r, column=1, value=metin)
        c.font, c.fill = baslik_font, baslik_fill
        r += 1

    def deger_satiri(etiket, deger, sayi=False):
        nonlocal r
        ws.cell(row=r, column=1, value=etiket).font = f10
        c = ws.cell(row=r, column=2, value=deger)
        c.font = f10
        if sayi:
            c.number_format = num_fmt
        r += 1

    # ── GENEL ──
    blok_baslik("GENEL")
    deger_satiri("Fatura Adedi", o["genel"]["adet"])
    for pb in sorted(o["genel"]["tutar"]):
        deger_satiri(f"Toplam Tutar ({pb})", o["genel"]["tutar"][pb], sayi=True)
    for pb in sorted(o["genel"]["kdv"]):
        deger_satiri(f"Toplam KDV ({pb})", o["genel"]["kdv"][pb], sayi=True)
    for kaynak in sorted(o["genel"]["kaynak"]):
        deger_satiri(f"Kaynak: {kaynak}", o["genel"]["kaynak"][kaynak])
    r += 1

    def kirilim_tablosu(baslik, ilk_kolon_adi, kayitlar):
        """kayitlar: [(ad, {"adet", "tutar": {pb: toplam}})]"""
        nonlocal r
        blok_baslik(baslik)
        for col, h in enumerate(
                [ilk_kolon_adi, "Adet", "Para Birimi", "Toplam Tutar"], 1):
            ws.cell(row=r, column=col, value=h).font = fbold
        r += 1
        for ad, v in kayitlar:
            pb_listesi = sorted(v["tutar"]) or [None]
            for i, pb in enumerate(pb_listesi):
                if i == 0:
                    ws.cell(row=r, column=1, value=ad).font = f10
                    ws.cell(row=r, column=2, value=v["adet"]).font = f10
                if pb is not None:
                    ws.cell(row=r, column=3, value=pb).font = f10
                    c = ws.cell(row=r, column=4, value=v["tutar"][pb])
                    c.font, c.number_format = f10, num_fmt
                r += 1
        r += 1

    kirilim_tablosu("AYLIK", "Ay", o["aylik"])
    kirilim_tablosu("ŞİRKET", "Şirket", o["sirket"])

    for col, w in [(1, 34), (2, 12), (3, 12), (4, 16)]:
        ws.column_dimensions[get_column_letter(col)].width = w
